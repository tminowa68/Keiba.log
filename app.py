# import openpyxl, re, os, json, glob
# from datetime import date, datetime, timedelta
# from flask import Flask, flash, render_template, request, redirect

# app = Flask(__name__)
# app.secret_key = 'your_secret_key_here'
# horse_data = '2026_horse_data.xlsx'

# def kana_to_hira(text):
#     return "".join([chr(ord(c) - 96) if "ァ" <= c <= "ヶ" else c for c in text])

# if not os.path.exists(horse_data):
#     wb = openpyxl.Workbook()

#     ws1 = wb.active
#     ws1.title = "Horses"
#     ws1.append(["馬名", "性別", "生年月日", "父", "母", "母父", "所属", "厩舎"])

#     ws_miho = wb.create_sheet(title="美浦")
#     ws_miho.append(["厩舎名", "よみがな"])

#     ws_ritto = wb.create_sheet(title="栗東")
#     ws_ritto.append(["厩舎名", "よみがな"])

#     wb.save(horse_data)

# def get_all_horses():
#     wb = openpyxl.load_workbook(horse_data, data_only=True)
#     return [list(row) for row in wb["Horses"].iter_rows(values_only=True)][1:]

# def get_stables_list():
#     wb = openpyxl.load_workbook(horse_data, data_only=True)
#     all_stables = []

#     for sheet_name in ["美浦", "栗東"]:
#         sheet = wb[sheet_name]
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             if row[0] and row[1]:
#                 all_stables.append({
#                     'display_name': f"{sheet_name}・{row[0]}", 
#                     'kana': row[1],
#                     'area': sheet_name
#                 })
#     return all_stables

# def calc_added_prize(rank, condition, race_name):
#     condition = str(condition or "")
#     race_name = str(race_name or "")
    
#     try:
#         rank = int(rank)
#     except (ValueError, TypeError):
#         return 0

#     # 1. 2歳重賞(G1,G2除く)の特例
#     if "重賞" in condition and ("2歳" in condition or "2歳" in race_name):
#         if "G1" not in race_name and "G2" not in race_name:
#             if rank == 1: return 1600
#             if rank == 2: return 600

#     # 2. それ以外は1着のみ加算
#     if rank != 1:
#         return 0

#     if "オープン" in condition:
#         is_listed = "L" in race_name or "リステッド" in race_name
#         if "2歳" in condition: return 800 if is_listed else 600
#         if "3歳" in condition: return 1200 if is_listed else 1000
#         return 1400 if is_listed else 1200
        
#     if "3勝" in condition or "1600万" in condition: return 900
#     if "2勝" in condition or "1000万" in condition: return 600
#     if "1勝" in condition or "500万" in condition: return 400
#     if "新馬" in condition or "未勝利" in condition: return 400
    
#     return 0

# def judge_class_by_prize(total_prize, has_raced=False):
#     if total_prize == 0: return "未勝利" if has_raced else "新馬"
#     if total_prize <= 500: return "1勝クラス"
#     if total_prize <= 1000: return "2勝クラス"
#     if total_prize <= 1600: return "3勝クラス"
#     return "オープン"

# def judge_required_class(condition_str):
#     """レース条件文字列から該当するクラス名を抽出"""
#     s = str(condition_str or "")
#     if "1勝" in s or "500万" in s: return "1勝クラス"
#     if "2勝" in s or "1000万" in s: return "2勝クラス"
#     if "3勝" in s or "1600万" in s: return "3勝クラス"
#     if "オープン" in s or "重賞" in s: return "オープン"
#     if "未勝利" in s: return "未勝利"
#     if "新馬" in s: return "新馬"
#     return None

# def get_class_from_results(horse_name, target_date_str, all_results_dict):
#     """
#     メモリ上の辞書(all_results_dict)から指定日前の賞金を計算する
#     """
#     total_prize = 0
#     target_dt = datetime.strptime(target_date_str, '%Y-%m-%d')
    
#     results = all_results_dict.get(horse_name, [])
#     past_races = []
    
#     for r in results:
#         # 日付比較
#         if datetime.strptime(r['date'], '%Y-%m-%d') < target_dt:
#             past_races.append(r)
#             total_prize += calc_added_prize(r['rank'], r['condition'], r['race_name'])
            
#     # 過去のレース出走数（len(past_races) > 0）を渡して判定
#     return judge_class_by_prize(total_prize, len(past_races) > 0)

# def get_horse_history(horse_name):
#     """
#     /horse/<name> のロジックを流用し、
#     特定の馬の全レース履歴をリストで返す
#     """
#     horse_races = []
#     entry_files = glob.glob('*_entry_*.xlsx')
#     year_master_wbs = {}

#     for f_path in entry_files:
#         filename = os.path.basename(f_path)
#         match = re.search(r'^(\d{4})_entry_(.+)\.xlsx$', filename)
#         if not match: continue
        
#         target_year = match.group(1)
#         race_data_file = f"{target_year}_race_data.xlsx"

#         if target_year not in year_master_wbs:
#             if os.path.exists(race_data_file):
#                 year_master_wbs[target_year] = openpyxl.load_workbook(race_data_file, data_only=True)
#             else:
#                 year_master_wbs[target_year] = None
        
#         race_master_wb = year_master_wbs[target_year]
#         if not race_master_wb: continue

#         # --- date_map / venue_map 生成 (ここはご提示のコードと同じ) ---
#         date_map = {}
#         if "schedule" in race_master_wb.sheetnames:
#             ws_sched = race_master_wb["schedule"]
#             for row in ws_sched.iter_rows(min_row=2, values_only=True):
#                 d_val = row[0]
#                 if d_val:
#                     d_str = d_val.strftime('%Y-%m-%d') if isinstance(d_val, datetime) else str(d_val)[:10]
#                     for i in [1, 4, 7]:
#                         if i+2 < len(row) and row[i+1]:
#                             date_map[f"{row[i]}回{row[i+1]}{row[i+2]}日"] = d_str

#         wb_e = openpyxl.load_workbook(f_path, data_only=True)
#         for sheet_name in wb_e.sheetnames:
#             ws_e = wb_e[sheet_name]
#             race_date = date_map.get(sheet_name, "1900-01-01") # 日付不明時は過去扱い

#             # 1R〜12R走査
#             for r_num in range(1, 13):
#                 name_col = (r_num - 1) * 4 + 3
#                 rank_col = (r_num - 1) * 4 + 1
#                 # race_masterから詳細を取るロジック（省略していますが、ご提示のコードをここに含めます）
#                 # ここでは簡易的に rank と condition が取得できている前提
#                 for r in range(3, ws_e.max_row + 1):
#                     if ws_e.cell(row=r, column=name_col).value == horse_name:
#                         # ここで calc_added_prize に必要な情報を収集
#                         # ※実際にはご提示のコードのように race_info_cache を使って
#                         # condition や race_name を取得してください
#                         res = {
#                             'date': race_date,
#                             'rank': ws_e.cell(row=r, column=rank_col).value,
#                             'condition': "仮の条件", # 実際には race_info_cache から取得
#                             'race_name': "仮のレース名" # 実際には race_info_cache から取得
#                         }
#                         horse_races.append(res)
#     return horse_races

# def load_all_horse_results():
#     """
#     全ての entry ファイルを走査し、全馬の成績を一つの辞書にまとめる。
#     戻り値: { "馬名": [ {"date": "2025-01-01", "rank": 1, "condition": "...", "name": "..."}, ... ], ... }
#     """
#     all_results = {}
#     entry_files = glob.glob('*_entry_*.xlsx')
    
#     # 年度ごとの race_data をキャッシュ（開く回数を最小限にする）
#     year_masters = {}

#     for f_path in entry_files:
#         filename = os.path.basename(f_path)
#         m = re.search(r'^(\d{4})_entry_(.+)\.xlsx$', filename)
#         if not m: continue
#         year = m.group(1)

#         # その年の race_master を1回だけ読み込む
#         if year not in year_masters:
#             master_path = f"{year}_race_data.xlsx"
#             if os.path.exists(master_path):
#                 # read_only=True にすると読み込みが非常に速くなります
#                 year_masters[year] = openpyxl.load_workbook(master_path, data_only=True, read_only=True)
        
#         wb_m = year_masters.get(year)
#         wb_e = openpyxl.load_workbook(f_path, data_only=True, read_only=True)

#         for sheet_name in wb_e.sheetnames:
#             ws_e = wb_e[sheet_name]
#             # ここで schedule シート等から日付(race_date)を取得するロジック（既存のものを流用）
#             # 今回は簡易的に変数 race_date とします
#             race_date = "2025-01-01" # 実際には取得ロジックを入れる

#             # 1R〜12Rを走査
#             for r_num in range(1, 13):
#                 name_col = (r_num - 1) * 4 + 3
#                 rank_col = (r_num - 1) * 4 + 1
                
#                 # 3行目から馬名をチェック
#                 for r in range(3, ws_e.max_row + 1):
#                     h_name = ws_e.cell(row=r, column=name_col).value
#                     if not h_name: continue
                    
#                     rank = ws_e.cell(row=r, column=rank_col).value
#                     # 成績データを辞書に追加
#                     if h_name not in all_results:
#                         all_results[h_name] = []
                    
#                     all_results[h_name].append({
#                         'date': race_date,
#                         'rank': rank,
#                         'condition': "レース条件", # ここも master から取得
#                         'race_name': "レース名"    # ここも master から取得
#                     })
#     return all_results

# @app.route('/')
# def index():
#     # フィルタリング条件（厩舎名など）
#     stable_filter = request.args.get('stable', '')
#     all_horses = get_all_horses()
#     results = all_horses
    
#     if stable_filter:
#         results = [h for h in results if h[6] == stable_filter]
        
#     current_year = datetime.now().year
#     return render_template('index.html', 
#                            results=results, 
#                            stables=get_stables_list(), 
#                            current_year=current_year)

# @app.route('/add_horse_page')
# def add_horse_page():
#     default_year = datetime.now().year - 3
#     # 厩舎リストが必要なので get_stables_list() を渡す
#     return render_template('add_horse.html',
#                            stables=get_stables_list(),
#                            default_year=default_year)

# @app.route('/add_horse', methods=['POST'])
# def add_horse():
#     wb = openpyxl.load_workbook(horse_data)
#     ws = wb["Horses"]

#     # フォームからデータを取得（リダイレクト用にnameを変数に入れる）
#     name = request.form.get('name')
#     year = int(request.form.get('year'))
#     month = int(request.form.get('month'))
#     day = int(request.form.get('day'))
#     birth_date = date(year, month, day)

#     for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
#         if row[0] == name:
#             flash(f"エラー：『{name}』は既に登録されています。")
#             return redirect('/add_horse_page')
        
#     # フォームから送られたデータを保存
#     wb["Horses"].append([
#         request.form.get('name'), 
#         request.form.get('gender'), 
#         birth_date,
#         request.form.get('sire'),
#         request.form.get('dam'),
#         request.form.get('area'),
#         request.form.get('stable_name')
#     ])

#     # --- テーブル機能への対応 ---
#     # そのシートにある最初のテーブルを取得
#     if ws.tables:
#         table_name = list(ws.tables.keys())[0]
#         table = ws.tables[table_name]

#         new_range = f"A1:G{ws.max_row}"
#         table.ref = new_range
    
#     # --- 並べ替え（ソート） ---
#     data_rows = []
#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
#         if row[0]: # 馬名がある場合
#             data_rows.append(row)

#     # 馬名（0列目）でソート
#     data_rows.sort(key=lambda x: x[0] if x[0] else "")

#     # 2行目以降を書き戻す
#     ws.delete_rows(2, ws.max_row)
#     for sorted_row in data_rows:
#         ws.append(sorted_row)
#         if sorted_row[2]:
#             ws.cell(row=ws.max_row, column=3).number_format = 'yyyy/m/d'

#     wb.save(horse_data)
#     return redirect(f"/horse/{name}")

# @app.route('/add_stable', methods=['POST'])
# def add_stable():
#     name = request.form.get('stable_name')
#     kana = request.form.get('kana')
#     area = request.form.get('area')
    
#     if name and area and area in ["美浦", "栗東"]:
#         wb = openpyxl.load_workbook(horse_data)
#         sheet = wb[area]
        
#         existing_stables = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
#         if name in existing_stables:
#             flash(f"エラー: {name}厩舎は既に{area}に登録されています。")
#             return redirect('/add_horse_page')

#         kana = re.sub(r'\s+', '', kana)
#         if 'kana_to_hira' in globals():
#             kana = kana_to_hira(kana)
            
#         sheet.append([name, kana])

#         # --- テーブル機能への対応 ---
#         # そのシートにある最初のテーブルを取得
#         if sheet.tables:
#             table_name = list(sheet.tables.keys())[0]
#             table = sheet.tables[table_name]

#             new_range = f"A1:E{sheet.max_row}"
#             table.ref = new_range
            
#             # --- 並べ替え（ソート） ---
#             data_rows = []
#             for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
#                 if row[0]:
#                     data_rows.append(row)

#             data_rows.sort(key=lambda x: x[1] if x[1] else "")

#             # 2行目以降を書き戻す
#             sheet.delete_rows(2, sheet.max_row)
#             for sorted_row in data_rows:
#                 sheet.append(sorted_row)

#         wb.save(horse_data)   
#     return redirect('/add_horse_page')

# @app.route('/horse/<name>')
# def horse_detail(name):
#     all_horses = get_all_horses()
#     current_year = datetime.now().year
#     horse = next((h for h in all_horses if h[0] == name), None)

#     # --- レース履歴の抽出ロジック ---
#     horse_races = []
#     entry_files = glob.glob('*_entry_*.xlsx')

#     # 年度ごとに race_master_wb をキャッシュする辞書
#     year_master_wbs = {}

#     for f_path in entry_files:
#         # ファイル名から年度と開催地を抽出 (例: 2025_entry_中山.xlsx)
#         filename = os.path.basename(f_path)
#         match = re.search(r'^(\d{4})_entry_(.+)\.xlsx$', filename)
#         if not match: continue
        
#         target_year = match.group(1)
#         venue = match.group(2)
#         race_data = f"{target_year}_race_data.xlsx"

#         # その年の race_master をまだ読み込んでいなければ読み込む
#         if target_year not in year_master_wbs:
#             if os.path.exists(race_data):
#                 year_master_wbs[target_year] = openpyxl.load_workbook(race_data, data_only=True)
#             else:
#                 year_master_wbs[target_year] = None
        
#         race_master_wb = year_master_wbs[target_year]
#         if not race_master_wb: continue

#         # --- その年専用の date_map と venue_map を生成 ---
#         date_map = {}
#         venue_map = {}
#         if "schedule" in race_master_wb.sheetnames:
#             ws_sched = race_master_wb["schedule"]
#             for row in ws_sched.iter_rows(min_row=2, values_only=True):
#                 d_val = row[0]
#                 if d_val:
#                     # 日付を YYYY-MM-DD 形式の文字列に変換
#                     d_str = d_val.strftime('%Y-%m-%d') if isinstance(d_val, datetime) else str(d_val)[:10]
#                     # 開催情報は 1, 4, 7 番目のインデックスから始まる
#                     for i in [1, 4, 7]:
#                         if i + 2 < len(row):
#                             v_id, v_name, v_day = row[i], row[i+1], row[i+2]
#                             if v_name:
#                                 # 例: "1回中山1日"
#                                 label = f"{v_id}回{v_name}{v_day}日"
#                                 date_map[label] = d_str
#                                 venue_map[label] = f"{v_id}_{v_name}"
#                 pass

#         wb_e = openpyxl.load_workbook(f_path, data_only=True)

#         for sheet_name in wb_e.sheetnames:
#             ws_e = wb_e[sheet_name]
            
#             race_date = date_map.get(sheet_name, sheet_name)
            
#             # --- 曜日の付与 ---
#             display_date = race_date
#             try:
#                 d_obj = datetime.strptime(race_date, '%Y-%m-%d')
#                 weekdays = ['月', '火', '水', '木', '金', '土', '日']
#                 display_date = f"{d_obj.strftime('%Y年%m月%d日')}({weekdays[d_obj.weekday()]})"
#             except ValueError:
#                 pass

#             # --- race_master からレース情報を一括抽出 ---
#             m_sheet_name = venue_map.get(sheet_name)
#             ws_m = race_master_wb[m_sheet_name] if m_sheet_name and m_sheet_name in race_master_wb.sheetnames else None
            
#             race_info_cache = {}
#             if ws_m:
#                 target_col_idx = next((cell.column for cell in ws_m[1] if cell.value == sheet_name), None)
#                 if target_col_idx:
#                     data_col_idx = target_col_idx 
#                     label_col_idx = target_col_idx - 1 

#                     max_r = ws_m.max_row
#                     last_row = 3
#                     for r in range(max_r, 2, -1):
#                         if ws_m.cell(row=r, column=data_col_idx + 1).value is not None:
#                             last_row = r
#                             break
#                     all_rows = list(ws_m.iter_rows(min_row=3, max_row=last_row))
#                     race_start_indices = {}

#                     # 各レースの開始位置を特定
#                     for i, row_cells in enumerate(all_rows):
#                         if label_col_idx < len(row_cells):
#                             val = str(row_cells[label_col_idx].value or "")
#                             match = re.search(r'(\d+)', val)
#                             if match:
#                                 r_num_m = int(match.group(1))
#                                 if 1 <= r_num_m <= 12 and r_num_m not in race_start_indices:
#                                     race_start_indices[r_num_m] = i

#                     # 抽出処理
#                     sorted_races = sorted(race_start_indices.items())
#                     for idx, (r_num_m, start_idx) in enumerate(sorted_races):
#                         if idx + 1 < len(sorted_races):
#                             end_idx = sorted_races[idx + 1][1] - 1
#                         else:
#                             end_idx = len(all_rows) - 1

#                         race_col_data = [str(all_rows[i][data_col_idx].value or "") for i in range(start_idx, end_idx + 1)]

#                         race_info_cache[r_num_m] = {
#                             'name': extract_race_name(race_col_data),
#                             'condition': extract_race_condition(race_col_data),
#                             'course': extract_race_course(race_col_data)
#                         }

#             # 1R〜12Rまで走査
#             for r_num in range(1, 13):
#                 base_col = (r_num - 1) * 4 + 1
#                 rank_col = base_col      # 1列目: 着順
#                 num_col  = base_col + 1  # 2列目: 馬番
#                 name_col = base_col + 2  # 3列目: 馬名
#                 stat_col = base_col + 3  # 4列目: 出走ステータス

#                 for row in range(3, ws_e.max_row + 1):
#                     cell_value = ws_e.cell(row=row, column=name_col).value
#                     if cell_value == name:
#                         status = ws_e.cell(row=row, column=stat_col).value
#                         rank = ws_e.cell(row=row, column=rank_col).value
                        
#                         r_name = race_info_cache.get(r_num, {}).get('name', '-')
#                         r_condition = race_info_cache.get(r_num, {}).get('condition', '-')
#                         r_course = race_info_cache.get(r_num, {}).get('course', '-')

#                         horse_races.append({
#                             'sort_date': race_date, # ソート用のオリジナル日付
#                             'date_label': display_date, # 表示用の曜日付き日付
#                             'venue': venue,
#                             'num': r_num,
#                             'name': r_name,
#                             'condition': r_condition,
#                             'course': r_course,
#                             'status': status if status else "-",
#                             'rank': rank if rank else "-"
#                         })

#     # 実日付 (YYYY-MM-DD形式) でソートされるようになるため、降順に正しく並びます
#     horse_races.sort(key=lambda x: (x['sort_date'], x['num']), reverse=True)

#     sire_info = {"line1": "不明", "line2": "不明"}
        
#     if horse and horse[3]:  # horse[3] は「父」の名前
#         try:
#             wb = openpyxl.load_workbook(horse_data, data_only=True)
#             if "Sire" in wb.sheetnames:
#                 ws_sire = wb["Sire"]
#                 # 2行目から探索 (1列目が父馬の名前と一致するか)
#                 for row in ws_sire.iter_rows(min_row=2, values_only=True):
#                     if row[0] == horse[3]:
#                         # 3列目をline1、4列目をline2に格納
#                         sire_info["line1"] = row[2] if row[2] else "不明"
#                         sire_info["line2"] = row[3] if row[3] else "不明"
#                         break
#         except Exception as e:
#             print(f"Error reading Sire sheet: {e}")
        
#     dam_info = {"line3": "不明", "line4": "不明"}
        
#     if horse and horse[4]:
#         try:
#             wb = openpyxl.load_workbook(horse_data, data_only=True)
#             if "Dam" in wb.sheetnames:
#                 ws_dam = wb["Dam"]
#                 for row in ws_dam.iter_rows(min_row=2, values_only=True):
#                     if row[0] == horse[4]:
#                         dam_info["line3"] = row[2] if row[2] else "不明"
#                         dam_info["line4"] = row[3] if row[3] else "不明"
#                         break
#         except Exception as e:
#             print(f"Error reading Dam sheet: {e}")

#     return render_template('horse_detail.html',
#                             horse=horse,
#                             horse_races=horse_races,
#                             sire_info=sire_info,
#                             dam_info=dam_info,
#                             current_year=current_year)

# @app.route('/edit_horse/<name>')
# def edit_horse(name):
#     all_horses = get_all_horses()
#     current_year = datetime.now().year
#     horse = next((h for h in all_horses if h[0] == name), None)
#     if not horse:
#         return "Horse not found", 404
    
#     stables = get_stables_list()

#     return render_template('edit_horse.html',
#                            horse=horse,
#                            stables=stables,
#                            current_year=current_year)

# # ----------------------------------------------------
# # 抽出用ヘルパー関数
# # ----------------------------------------------------

# def format_excel_time(time_val):
#     if time_val is None: return ""
#     if isinstance(time_val, (int, float)):
#         return (datetime(2000, 1, 1) + timedelta(days=float(time_val))).strftime('%H:%M')
#     return time_val.strftime('%H:%M') if isinstance(time_val, datetime) else str(time_val)

# # レース名を抽出（最後から2行分を除いた部分）
# def extract_race_name(race_col_data):
#     # レース名がない場合（条件とコースの2行しかない場合など）
#     if len(race_col_data) <= 2:
#         return ""

#     # 最後から2行（条件とコース）を除外した配列がレース名の対象
#     name_strs = race_col_data[:-2]
#     hit_text = ""

#     # G(グレード)またはL(リステッド)、J・Gを含むセルを優先
#     for cell in name_strs:
#         cell_str = str(cell)
#         if "G" in cell_str or "L" in cell_str or "J・G" in cell_str:
#             hit_text = cell_str
#             break

#     # 見つからない場合は、名前ブロックの一番下（条件の1つ上）を採用
#     if not hit_text and name_strs:
#         hit_text = str(name_strs[-1])

#     # 「第〇〇回」などの余計な文字を削除
#     formatted_text = re.sub(r'第\s*\d+\s*回', '', hit_text).strip()
#     formatted_text = re.sub(r'第.*?回\s*', '', formatted_text).strip()
#     return formatted_text

# # 出走条件を抽出（下から2行目）
# def extract_race_condition(race_col_data):
#     if len(race_col_data) < 2:
#         return ""
    
#     cond_str = str(race_col_data[-2])

#     # --- 馬齢の抽出ロジック ---
#     age_part = ""
#     match_age = re.search(r'(\d歳(?:以上)?)', cond_str)
#     if match_age:
#         age_part = match_age.group(1)

#     # --- クラスの抽出ロジック ---
#     class_part = ""
#     if "障害" in cond_str:
#         class_part = "障害"
    
#     if "未勝利" in cond_str:
#         class_part += "未勝利"
#     elif "新馬" in cond_str:
#         class_part += "新馬"
#     elif "1勝クラス" in cond_str:
#         class_part += "1勝クラス"
#     elif "2勝クラス" in cond_str:
#         class_part += "2勝クラス"
#     elif "3勝クラス" in cond_str:
#         class_part += "3勝クラス"
#     elif "オープン" in cond_str:
#         class_part += "オープン"

#     return f"{age_part} {class_part}".strip()

# # コース情報を抽出（一番下の行）
# def extract_race_course(race_col_data):
#     if not race_col_data:
#         return ""
        
#     course_str = str(race_col_data[-1])
#     cond_str = str(race_col_data[-2]) if len(race_col_data) > 1 else ""

#     # --- コースの抽出ロジック ---
#     cource_part = ""
#     # 条件名に「障害」と入っているか、コースに「障」があれば「障」
#     if "障害" in cond_str or "障" in course_str:
#         cource_part = "障"
#     elif "芝" in course_str:
#         cource_part = "芝"
#     elif "ダ" in course_str:
#         cource_part = "ダ"

#     straight_part = "直線 " if "直" in course_str else ""

#     # --- 距離の抽出ロジック（数字とカンマを抽出） ---
#     dist_match = re.search(r'[\d,]+', course_str)
#     distance_part = dist_match.group(0) if dist_match else ""

#     # f"{cource_part} {straight_part} {distance_part}m" の形式で結合
#     # ※元のコード出力の空白感を維持するためそのままにしています
#     return f"{cource_part} {straight_part} {distance_part}m".strip()

# # ----------------------------------------------------
# # メインのルーティング関数
# # ----------------------------------------------------

# @app.route('/races')
# def race_list():
#     req_date = request.args.get('date')
#     if req_date:
#         target_year = req_date[:4]
#     else:
#         target_year = str(datetime.now().year)
    
#     race_data = f"{target_year}_race_data.xlsx"

#     available_dates = []
#     venue_data_map = {} 
#     day_races = {i: None for i in range(1, 13)}
#     search_text = "開催情報が見つかりません"

#     if os.path.exists(race_data):
#         try:
#             wb = openpyxl.load_workbook(race_data, data_only=True, read_only=True)
#             # --- scheduleシートの解析 ---
#             if "schedule" in wb.sheetnames:
#                 ws_sched = wb["schedule"]
#                 for row in ws_sched.iter_rows(min_row=2, values_only=True):
#                     d_val = row[0]
#                     if d_val:
#                         d_str = d_val.strftime('%Y-%m-%d') if isinstance(d_val, datetime) else str(d_val)
#                         available_dates.append(d_str)
#                         v_info = {}
#                         for i in [1, 4, 7]:
#                             v_id, v_name, v_day = row[i], row[i+1], row[i+2]
#                             if v_name:
#                                 v_info[str(v_name)] = {"id": v_id, "day": v_day}
#                         venue_data_map[d_str] = v_info
#         except PermissionError:
#             return "Excelファイルが開かれています。閉じてから再試行してください。"

#     # --- 2. デフォルト値の決定 ---
#     today_str = datetime.now().strftime('%Y-%m-%d')
#     future_dates = [d for d in available_dates if d >= today_str]
#     date = request.args.get('date', future_dates[0] if future_dates else (available_dates[-1] if available_dates else today_str))

#     current_day_venues = venue_data_map.get(date, {})
#     default_venue = list(current_day_venues.keys())[0] if current_day_venues else ''
#     venue = request.args.get('venue', default_venue)
    
#     # 日付ブロック表示用の処理
#     weekdays = ['月', '火', '水', '木', '金', '土', '日']
#     sorted_dates = sorted(available_dates)
#     date_blocks = []
#     if sorted_dates:
#         current_block = [sorted_dates[0]]
#         for i in range(1, len(sorted_dates)):
#             d1, d2 = datetime.strptime(sorted_dates[i-1], '%Y-%m-%d'), datetime.strptime(sorted_dates[i], '%Y-%m-%d')
#             if (d2 - d1).days <= 2: current_block.append(sorted_dates[i])
#             else: date_blocks.append(current_block); current_block = [sorted_dates[i]]
#         date_blocks.append(current_block)

#     current_date_block = next((b for b in date_blocks if date in b), [])
#     display_dates = [{'value': d, 'label': f"{datetime.strptime(d, '%Y-%m-%d').month}/{datetime.strptime(d, '%Y-%m-%d').day}({weekdays[datetime.strptime(d, '%Y-%m-%d').weekday()]})"} for d in current_date_block]

#     # --- 5. 対象シートからレース情報の抽出 ---
#     venue_info = current_day_venues.get(venue)
#     if venue_info:
#         target_sheet_name = f"{venue_info['id']}_{venue}"
#         if target_sheet_name in wb.sheetnames:
#             ws_race = wb[target_sheet_name]
#             search_text = f"{venue_info['id']}回{venue}{venue_info['day']}日"

#             # 対象列の特定
#             target_col_idx = next((cell.column for cell in ws_race[1] if cell.value == search_text), None)
#             if target_col_idx:
#                 data_col_idx = target_col_idx 
#                 label_col_idx = target_col_idx - 1 # 0始まりの配列に合わせる調整

#                 max_r = ws_race.max_row
#                 last_row = 3
#                 for r in range(max_r, 2, -1):
#                     if ws_race.cell(row=r, column=data_col_idx + 1).value is not None:
#                         last_row = r
#                         break
#                 all_rows = list(ws_race.iter_rows(min_row=3, max_row=last_row))
#                 race_start_indices = {}

#                 # 各レースの開始位置を特定
#                 for i, row_cells in enumerate(all_rows):
#                     if label_col_idx < len(row_cells):
#                         val = str(row_cells[label_col_idx].value or "")
#                         match = re.search(r'(\d+)', val)
#                         if match:
#                             r_num = int(match.group(1))
#                             if 1 <= r_num <= 12 and r_num not in race_start_indices:
#                                 race_start_indices[r_num] = i

#                 # 抽出処理
#                 sorted_races = sorted(race_start_indices.items())
#                 for idx, (r_num, start_idx) in enumerate(sorted_races):
                    
#                     # 終点の決定
#                     if idx + 1 < len(sorted_races):
#                         next_start_idx = sorted_races[idx + 1][1]
#                         end_idx = next_start_idx - 1
#                     else:
#                         # 12レースは末尾まで
#                         end_idx = len(all_rows) - 1

#                     race_col_data = [str(all_rows[i][data_col_idx].value or "") for i in range(start_idx, end_idx + 1)]

#                     day_races[r_num] = {
#                         'time': format_excel_time(all_rows[start_idx][data_col_idx +1].value),
#                         'num': all_rows[start_idx][label_col_idx].value,
#                         'name': extract_race_name(race_col_data),
#                         'condition': extract_race_condition(race_col_data),
#                         'course': extract_race_course(race_col_data)
#                     }

#     return render_template('race_list.html',
#                            date=date,
#                            venue=venue,
#                            day_races=day_races,
#                            available_venues=list(current_day_venues.keys()),
#                            available_dates=available_dates,
#                            display_dates=display_dates,
#                            search_text=search_text)

# @app.route('/edit_race')
# def race_detail():
#     req_date = request.args.get('date')
#     if req_date:
#         target_year = req_date[:4]
#     else:
#         target_year = str(datetime.now().year)
    
#     race_data = f"{target_year}_race_data.xlsx"

#     r_num_target = request.args.get('num', default=1, type=int)
#     available_dates = []
#     venue_data_map = {} 
#     target_race_data = None
#     search_text = "開催情報が見つかりません"
    
#     if os.path.exists(race_data):
#         try:
#             wb = openpyxl.load_workbook(race_data, data_only=True, read_only=True)
#             # --- scheduleシートの解析 ---
#             if "schedule" in wb.sheetnames:
#                 ws_sched = wb["schedule"]
#                 for row in ws_sched.iter_rows(min_row=2, values_only=True):
#                     d_val = row[0]
#                     if d_val:
#                         d_str = d_val.strftime('%Y-%m-%d') if isinstance(d_val, datetime) else str(d_val)
#                         available_dates.append(d_str)
#                         v_info = {}
#                         for i in [1, 4, 7]:
#                             v_id, v_name, v_day = row[i], row[i+1], row[i+2]
#                             if v_name:
#                                 v_info[str(v_name)] = {"id": v_id, "day": v_day}
#                         venue_data_map[d_str] = v_info
#         except PermissionError:
#             return "Excelファイルが開かれています。閉じてから再試行してください。"

#     # --- 2. デフォルト値の決定 ---
#     today_str = datetime.now().strftime('%Y-%m-%d')
#     future_dates = [d for d in available_dates if d >= today_str]
#     date = request.args.get('date', future_dates[0] if future_dates else (available_dates[-1] if available_dates else today_str))

#     current_day_venues = venue_data_map.get(date, {})
#     default_venue = list(current_day_venues.keys())[0] if current_day_venues else ''
#     venue = request.args.get('venue', default_venue)
    
#     # 日付ブロック表示用の処理
#     sorted_dates = sorted(available_dates)
#     date_blocks = []
#     if sorted_dates:
#         current_block = [sorted_dates[0]]
#         for i in range(1, len(sorted_dates)):
#             d1, d2 = datetime.strptime(sorted_dates[i-1], '%Y-%m-%d'), datetime.strptime(sorted_dates[i], '%Y-%m-%d')
#             if (d2 - d1).days <= 2: current_block.append(sorted_dates[i])
#             else: date_blocks.append(current_block); current_block = [sorted_dates[i]]
#         date_blocks.append(current_block)

#     # --- 5. 対象シートからレース情報の抽出 ---
#     venue_info = current_day_venues.get(venue)
#     if venue_info:
#         target_sheet_name = f"{venue_info['id']}_{venue}"
#         if target_sheet_name in wb.sheetnames:
#             ws_race = wb[target_sheet_name]
#             search_text = f"{venue_info['id']}回{venue}{venue_info['day']}日"

#             # 対象列の特定
#             target_col_idx = next((cell.column for cell in ws_race[1] if cell.value == search_text), None)
#             if target_col_idx:
#                 data_col_idx = target_col_idx 
#                 label_col_idx = target_col_idx - 1 # 0始まりの配列に合わせる調整

#                 max_r = ws_race.max_row
#                 last_row = 3
#                 for r in range(max_r, 2, -1):
#                     if ws_race.cell(row=r, column=data_col_idx + 1).value is not None:
#                         last_row = r
#                         break
#                 all_rows = list(ws_race.iter_rows(min_row=3, max_row=last_row))
#                 race_start_indices = {}

#                 # 各レースの開始位置を特定
#                 for i, row_cells in enumerate(all_rows):
#                     val = str(row_cells[label_col_idx].value or "")
#                     match = re.search(r'(\d+)', val)
#                     if match:
#                         num = int(match.group(1))
#                         if 1 <= num <= 12 and num not in race_start_indices:
#                             race_start_indices[num] = i

#                 # 抽出処理
#                 sorted_starts = sorted(race_start_indices.items())
#                 for idx, (num, start_idx) in enumerate(sorted_starts):
#                     # ターゲットのレース番号以外はスキップして処理を軽くする
#                     if num != r_num_target:
#                         continue

#                     # 終点の決定
#                     if idx + 1 < len(sorted_starts):
#                         end_idx = sorted_starts[idx + 1][1] - 1
#                     else:
#                         end_idx = len(all_rows) - 1

#                     race_col_data = [str(all_rows[i][data_col_idx].value or "") for i in range(start_idx, end_idx + 1)]

#                     target_race_data = {
#                         'time': format_excel_time(all_rows[start_idx][data_col_idx +1].value),
#                         'num': all_rows[start_idx][label_col_idx].value,
#                         'name': extract_race_name(race_col_data),
#                         'condition': extract_race_condition(race_col_data),
#                         'course': extract_race_course(race_col_data)
#                     }
#                     break # 見つかったらループ終了

#     # このレースが求めているクラスを判定
#     available_horses_with_class = []
    
#     if target_race_data:
#         # 1. 全馬の過去成績を「一度だけ」一括で読み込む（これが高速化のキモ）
#         all_results_dict = load_all_horse_results()

#         # 2. レースの要求クラスを判定
#         race_req_class = judge_required_class(target_race_data['condition']) if target_race_data else None

#         # 3. リスト作成（ファイルを開かず、メモリ内の辞書を検索するだけなので爆速）
#         all_horses_raw = get_all_horses()
#         available_horses_with_class = []

#         for h in all_horses_raw:
#             h_name = h[0]
#             h_gender = h[1]
#             birth_date = h[2]

#             # 馬齢の計算
#             try:
#                 if isinstance(birth_date, datetime):
#                     birth_year = birth_date.year
#                 else:
#                     birth_year = int(str(birth_date).split('/')[0])
                
#                 calculated_age = int(target_year) - birth_year
#             except (ValueError, TypeError, IndexError):
#                 calculated_age = "不明"

#             # メモリ内のデータでクラス判定
#             current_class = get_class_from_results(h_name, date, all_results_dict)
            
#             if race_req_class == "オープン" or current_class == race_req_class or (race_req_class == "未勝利" and current_class == "新馬"):
#                 available_horses_with_class.append({
#                     'name': h_name,
#                     'gender': h_gender,
#                     'age': calculated_age,
#                     'class': current_class,
#                     'stable': f"{h[5]}・{h[6]}"
#                 })
                
#     entered_horses = []
#     entry_file = f'{target_year}_entry_{venue}.xlsx'
#     if os.path.exists(entry_file):
#         wb_entry = openpyxl.load_workbook(entry_file, data_only=True)
#         if search_text in wb_entry.sheetnames:
#             ws_entry = wb_entry[search_text]
            
#             # 1レース4列構成 (着順, 馬番, 馬名, 出走)
#             rank_col = (int(r_num_target) - 1) * 4 + 1
#             num_col = rank_col + 1
#             name_col = rank_col + 2
#             status_col = rank_col + 3
            
#             for row in range(3, ws_entry.max_row + 1):
#                 rank = ws_entry.cell(row=row, column=rank_col).value
#                 num = ws_entry.cell(row=row, column=num_col).value
#                 name = ws_entry.cell(row=row, column=name_col).value
#                 status = ws_entry.cell(row=row, column=status_col).value
                
#                 if name:
#                     entered_horses.append({
#                         'rank': rank if rank else "",
#                         'num': num if num else "",
#                         'name': name, 
#                         'status': status
#                     })

#     return render_template('race_detail.html',
#                            date=date,
#                            venue=venue,
#                            race=target_race_data, # 単一のレースデータを渡す
#                            available_dates=available_dates,
#                            available_horses=available_horses_with_class,
#                            all_horses=all_horses_raw,
#                            entered_horses=entered_horses,
#                            search_text=search_text)

# @app.route('/save_entry', methods=['POST'])
# def save_entry():
#     data = request.json

#     # リクエストデータから日付(または年)を取得。なければ今年の西暦
#     date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
#     target_year = date_str[:4]

#     race_num = int(data.get('race_num'))
#     horse_name = data.get('horse_name')
#     entry_type = data.get('entry_type')
#     sheet_name = data.get('sheet_name')
    
#     # 馬番と着順（確定の時のみ送られてくる）
#     horse_num = data.get('horse_num')
#     horse_rank = data.get('horse_rank')

#     status_label = {"estimated": "想定", "special": "特別", "final": "確定"}.get(entry_type, "想定")
#     file_name = f'{target_year}_entry_{data.get("venue")}.xlsx'
    
#     if os.path.exists(file_name):
#         wb = openpyxl.load_workbook(file_name)
#     else:
#         wb = openpyxl.Workbook()
#         if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

#     if sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#     else:
#         ws = wb.create_sheet(title=sheet_name)
#         for r in range(1, 13):
#             start_col = (r - 1) * 4 + 1
#             ws.cell(row=1, column=start_col, value=f"{r}R")
#             ws.cell(row=2, column=start_col, value="着順")
#             ws.cell(row=2, column=start_col + 1, value="馬番")
#             ws.cell(row=2, column=start_col + 2, value="馬名")
#             ws.cell(row=2, column=start_col + 3, value="出走")

#     # save_entry 関数の列計算部分
#     base_col = (race_num - 1) * 4 + 1
#     rank_col = base_col
#     num_col = base_col + 1
#     name_col = base_col + 2
#     status_col = base_col + 3
    
#     # 既存行の検索（name_colを使用）
#     target_row = 3
#     found_row = None
#     while ws.cell(row=target_row, column=name_col).value is not None:
#         if ws.cell(row=target_row, column=name_col).value == horse_name:
#             found_row = target_row
#             break
#         target_row += 1
    
#     write_row = found_row if found_row else target_row
    
#     # データの書き込み
#     ws.cell(row=write_row, column=name_col, value=horse_name)
#     ws.cell(row=write_row, column=status_col, value=status_label)
    
#     # 確定データがある場合は数値として保存
#     if horse_num:
#         ws.cell(row=write_row, column=num_col, value=int(horse_num))
#     if horse_rank:
#         ws.cell(row=write_row, column=rank_col, value=int(horse_rank))

#     wb.save(file_name)
#     return {"status": "success"}, 200

# @app.route('/add_parent', methods=['GET', 'POST'])
# def add_parent():
#     if request.method == 'POST':
#         # 1. フォームデータの取得
#         origin = request.form.get('origin')  # 仔の名前（リダイレクト用）
#         p_type = request.form.get('p_type')  # "Sire" か "Dam"
#         p_name = request.form.get('p_name')
        
#         y = request.form.get('year')
#         m = request.form.get('month')
#         d = request.form.get('day')
#         sire = request.form.get('sire')
#         dam = request.form.get('dam')

#         # 2. 保存処理
#         try:
#             wb = openpyxl.load_workbook(horse_data)
#             sheet_name = p_type  # "Sire" または "Dam"
            
#             # シートが存在しない場合は作成
#             if sheet_name not in wb.sheetnames:
#                 ws = wb.create_sheet(title=sheet_name)
#                 ws.append(["馬名", "生年月日", "父", "母"])
#             else:
#                 ws = wb[sheet_name]

#             birth_date = None
#             if y and m and d:
#                 try:
#                     birth_date = date(int(y), int(m), int(d))
#                 except ValueError:
#                     # 2月31日などの不正な日付が入った場合の対策
#                     birth_date = None

#             # 既存データの更新か、新規追加かの判定
#             found = False
#             for row in ws.iter_rows(min_row=2):
#                 if row[0].value == p_name:
#                     row[1].value = birth_date
#                     row[2].value = sire
#                     row[3].value = dam
#                     found = True
#                     break
            
#             if not found:
#                 ws.append([p_name, birth_date, sire, dam])

#             # --- テーブル機能への対応 ---
#             # そのシートにある最初のテーブルを取得
#             if ws.tables:
#                 table_name = list(ws.tables.keys())[0]
#                 table = ws.tables[table_name]
                
#                 # データの最終範囲を計算（例: A1からD列の最終行まで）
#                 # ws.max_row が新しく追加された後の行数を指します
#                 new_range = f"A1:D{ws.max_row}"
#                 table.ref = new_range
            
#             # --- 並べ替え（ソート） ---
#             # テーブル機能を使っていても、データの並べ替えはPython側でリスト化して行うのが確実です
#             data_rows = []
#             for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
#                 if row[0]: # 馬名がある場合
#                     data_rows.append(row)

#             # 馬名（0列目）でソート
#             data_rows.sort(key=lambda x: x[0] if x[0] else "")

#             # 2行目以降を書き戻す
#             ws.delete_rows(2, ws.max_row)
#             for sorted_row in data_rows:
#                 ws.append(sorted_row)
#                 if sorted_row[1]:
#                     ws.cell(row=ws.max_row, column=2).number_format = 'yyyy/m/d'

#             wb.save(horse_data)

#         except Exception as e:
#             flash(f"エラーが発生しました: {e}")

#         return redirect(f"/horse/{origin}") if origin else redirect('/')

#     # GET時の処理
#     p_type = request.args.get('type', 'Sire')
#     p_name = request.args.get('name', '')
#     origin = request.args.get('origin', '')
    
#     # Excelから既存データを検索（デフォルト値用）
#     existing_data = {"year": "", "month": "", "day": "", "sire": "", "dam": ""}
#     sheet_name = "Sire" if p_type == "Sire" else "Dam"
    
#     try:
#         wb = openpyxl.load_workbook(horse_data, data_only=True)
#         if sheet_name in wb.sheetnames:
#             ws = wb[sheet_name]
#             for row in ws.iter_rows(values_only=True):
#                 if row[0] == p_name:
#                     # 生年月日の分解
#                     if row[1] and isinstance(row[1], (date, datetime)):
#                         existing_data["year"] = row[1].year
#                         existing_data["month"] = row[1].month
#                         existing_data["day"] = row[1].day
#                     existing_data["sire"] = row[2] or ""
#                     existing_data["dam"] = row[3] or ""
#                     break
#     except:
#         pass

#     return render_template('add_parent.html', 
#                            p_type=p_type,
#                            p_name=p_name,
#                            origin=origin,
#                            data=existing_data)

# @app.route('/update_horse', methods=['POST'])
# def update_horse():
#     old_name = request.form.get('old_name')
#     new_name = request.form.get('name')
#     gender = request.form.get('gender')
    
#     # HTMLの name="year", "month", "day" に合わせる
#     y = request.form.get('year')
#     m = request.form.get('month')
#     d = request.form.get('day')
    
#     # 厩舎情報（JSでセットされる隠しフィールド）
#     location = request.form.get('location')
#     stable_name = request.form.get('stable_name')

#     # 必須項目がNoneでないかチェック
#     if not all([y, m, d, new_name]):
#         return "入力データが不足しています（生年月日など）。", 400

#     try:
#         birth_date = date(int(y), int(m), int(d))
        
#         wb = openpyxl.load_workbook(horse_data)
#         ws = wb["Horses"]
        
#         for row in ws.iter_rows(min_row=2):
#             if row[0].value == old_name:
#                 row[0].value = new_name
#                 row[1].value = gender
#                 row[2].value = birth_date
#                 row[3].value = request.form.get('sire')
#                 row[4].value = request.form.get('dam')
#                 row[5].value = location
#                 row[6].value = stable_name
#                 break
        
#         wb.save(horse_data)
#         return redirect(f"/horse/{new_name}")

#     except ValueError as e:
#         return f"日付が正しくありません: {e}", 400
#     except Exception as e:
#         return f"エラーが発生しました: {e}", 500

# if __name__ == '__main__':
#     app.run(debug=True, port=5001 ,use_reloader=False)

import sqlite3
import openpyxl, re, os, json, glob
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.secret_key = 'your_secret_key_here' # セッション用の秘密鍵
DB_NAME = 'users.db'
horse_data = '2026_horse_data.xlsx'

# --- データベース初期化 ---
def add_new_user(username, password):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        hashed_pw = generate_password_hash(password)
        try:
            cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_pw))
            conn.commit()
            print(f"ユーザー {username} を登録しました。")
        except sqlite3.IntegrityError:
            print("そのユーザー名は既に存在します。")

# add_new_user('', '')

def delete_user(username):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE username = ?", (username,))
        conn.commit()
        if cursor.rowcount > 0:
            print(f"ユーザー {username} を削除しました。")
        else:
            print(f"ユーザー {username} が見つかりませんでした。")

# delete_user('')

# --- 認証用デコレータ ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('login'):
            flash('この操作にはログインが必要です。')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ==============================================================================
# 初期設定
# ==============================================================================
if not os.path.exists(horse_data):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Horses"
    ws1.append(["馬名", "性別", "生年月日", "父", "母", "母父", "所属", "厩舎"])
    ws_miho = wb.create_sheet(title="美浦")
    ws_miho.append(["厩舎名", "よみがな"])
    ws_ritto = wb.create_sheet(title="栗東")
    ws_ritto.append(["厩舎名", "よみがな"])
    wb.save(horse_data)

# ==============================================================================
# 共通ヘルパー関数群
# ==============================================================================

def kana_to_hira(text):
    return "".join([chr(ord(c) - 96) if "ァ" <= c <= "ヶ" else c for c in text])

def get_all_horses():
    wb = openpyxl.load_workbook(horse_data, data_only=True)
    return [list(row) for row in wb["Horses"].iter_rows(values_only=True)][1:]

def get_stables_list():
    wb = openpyxl.load_workbook(horse_data, data_only=True)
    all_stables = []
    for sheet_name in ["美浦", "栗東"]:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                all_stables.append({
                    'display_name': f"{sheet_name}・{row[0]}", 
                    'kana': row[1],
                    'area': sheet_name
                })
    return all_stables

# --- 賞金・クラス判定 ---
def calc_added_prize(rank, condition, race_name, horse_birthday, race_date_str):
    """
    レース結果から「収得賞金」を計算する
    """
    condition, race_name = str(condition or ""), str(race_name or "")
    try: rank = int(rank)
    except (ValueError, TypeError): return 0

    # レース当時の馬齢を計算（2025年以前の問題を解決するために重要）
    race_year = int(race_date_str[:4])
    birth_year = horse_birthday.year if isinstance(horse_birthday, datetime) else int(str(horse_birthday)[:4])
    age = race_year - birth_year

    # 収得賞金が加算されるのは基本的に2着以内（重賞）か1着のみ
    # 重賞（G1, G2, G3）の場合
    if "G" in race_name or "重賞" in condition or "J・G" in race_name:
        if rank == 1:
            if age <= 2: return 600
            if age == 3: return 1200 if race_date_str <= f"{race_year}-06-30" else 1200 # 3歳夏までは変動あり
            return 1200 # 古馬重賞1着
        if rank == 2:
            if age <= 2: return 200 # 2歳重賞2着
            if age == 3 and race_date_str <= f"{race_year}-06-30": return 400 # 3歳春重賞2着
            return 0 # 基本的に重賞2着で加算されるのは3歳春まで
    
    # 1着の場合の加算（一般戦）
    if rank == 1:
        if "新馬" in condition or "未勝利" in condition: return 400
        if "1勝" in condition or "500万" in condition: return 500
        if "2勝" in condition or "1000万" in condition: return 600
        if "3勝" in condition or "1600万" in condition: return 900
        if "オープン" in condition:
            is_listed = "L" in race_name or "リステッド" in race_name
            return 800 if is_listed else 700
            
    return 0

def judge_class_by_prize(total_prize, has_raced, age, race_month):
    if total_prize == 0: 
        return "未勝利" if has_raced else "新馬"
    
    # 2歳、または 5月までの3歳
    if age == 2 or (age == 3 and race_month <= 5):
        if total_prize <= 500: return "1勝クラス"
        return "オープン"
    # 6月以降の3歳、および 4歳以上
    else:
        if total_prize <= 500: return "1勝クラス"
        if total_prize <= 1000: return "2勝クラス"
        if total_prize <= 1600: return "3勝クラス"
        return "オープン"

def judge_required_class(condition_str):
    s = str(condition_str or "")
    if "1勝" in s or "500万" in s: return "1勝クラス"
    if "2勝" in s or "1000万" in s: return "2勝クラス"
    if "3勝" in s or "1600万" in s: return "3勝クラス"
    if "オープン" in s or "重賞" in s: return "オープン"
    if "未勝利" in s: return "未勝利"
    if "新馬" in s: return "新馬"
    return None

def get_class_from_results(horse_name, target_date_str, all_results_dict, horse_birthday):
    total_prize = 0
    target_dt = datetime.strptime(target_date_str, '%Y-%m-%d')
    results = all_results_dict.get(horse_name, [])
    
    # ターゲットの日付より前のレースのみを集計
    past_races = [r for r in results if datetime.strptime(r['date'], '%Y-%m-%d') < target_dt]
    
    for r in past_races:
        # 馬の誕生日を渡して、その時の年齢に基づいた賞金を計算
        total_prize += calc_added_prize(r['rank'], r['condition'], r['race_name'], horse_birthday, r['date'])
            
    # レース時点の馬齢と月を計算
    race_year = target_dt.year
    race_month = target_dt.month
    birth_year = horse_birthday.year if isinstance(horse_birthday, datetime) else int(str(horse_birthday)[:4])
    age = race_year - birth_year

    return judge_class_by_prize(total_prize, len(past_races) > 0, age, race_month)

def load_all_horse_results():
    all_results = {}
    entry_files = glob.glob('*_entry_*.xlsx')
    year_masters = {}

    for f_path in entry_files:
        m = re.search(r'^(\d{4})_entry_(.+)\.xlsx$', os.path.basename(f_path))
        if not m: continue
        year = m.group(1)

        if year not in year_masters:
            master_path = f"{year}_race_data.xlsx"
            if os.path.exists(master_path):
                year_masters[year] = openpyxl.load_workbook(master_path, data_only=True, read_only=True)
        
        wb_e = openpyxl.load_workbook(f_path, data_only=True, read_only=True)

        for sheet_name in wb_e.sheetnames:
            ws_e = wb_e[sheet_name]
            race_date = "2025-01-01" # 簡易仮置き。必要に応じて取得ロジックへ
            for r_num in range(1, 13):
                name_col, rank_col = (r_num - 1) * 4 + 3, (r_num - 1) * 4 + 1
                for r in range(3, ws_e.max_row + 1):
                    h_name = ws_e.cell(row=r, column=name_col).value
                    if not h_name: continue
                    if h_name not in all_results: all_results[h_name] = []
                    all_results[h_name].append({
                        'date': race_date,
                        'rank': ws_e.cell(row=r, column=rank_col).value,
                        'condition': "レース条件",
                        'race_name': "レース名"
                    })
    return all_results

# --- レース情報の抽出 ---
def format_excel_time(time_val):
    if time_val is None: return ""
    if isinstance(time_val, (int, float)):
        return (datetime(2000, 1, 1) + timedelta(days=float(time_val))).strftime('%H:%M')
    return time_val.strftime('%H:%M') if isinstance(time_val, datetime) else str(time_val)

def extract_race_name(race_col_data):
    if len(race_col_data) <= 2: return ""
    name_strs = race_col_data[:-2]
    hit_text = next((str(c) for c in name_strs if any(x in str(c) for x in ["G", "L", "J・G"])), None)
    if not hit_text and name_strs: hit_text = str(name_strs[-1])
    formatted_text = re.sub(r'第\s*\d+\s*回', '', hit_text).strip()
    return re.sub(r'第.*?回\s*', '', formatted_text).strip()

def extract_race_condition(race_col_data):
    if len(race_col_data) < 2: return ""
    cond_str = str(race_col_data[-2])
    match_age = re.search(r'(\d歳(?:以上)?)', cond_str)
    age_part = match_age.group(1) if match_age else ""
    class_part = "障害" if "障害" in cond_str else ""
    for c in ["未勝利", "新馬", "1勝クラス", "2勝クラス", "3勝クラス", "オープン"]:
        if c in cond_str: class_part += c
    return f"{age_part} {class_part}".strip()

def extract_race_course(race_col_data):
    if not race_col_data: return ""
    course_str = str(race_col_data[-1])
    cond_str = str(race_col_data[-2]) if len(race_col_data) > 1 else ""
    cource_part = "障" if "障害" in cond_str or "障" in course_str else ("芝" if "芝" in course_str else ("ダ" if "ダ" in course_str else ""))
    straight_part = "直線 " if "直" in course_str else ""
    dist_match = re.search(r'[\d,]+', course_str)
    distance_part = dist_match.group(0) if dist_match else ""
    return f"{cource_part} {straight_part} {distance_part}m".strip()

# --- ファイル操作・整理の共通処理 ---
def sort_and_resize_table(ws, sort_col_index=0, date_format_col=None):
    """Excelのテーブルをリサイズし、指定列でソートして上書きする共通関数"""
    if ws.tables:
        table_name = list(ws.tables.keys())[0]
        table = ws.tables[table_name]
        max_col_letter = get_column_letter(ws.max_column)
        table.ref = f"A1:{max_col_letter}{ws.max_row}"
        
    data_rows = [row for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True) if row[0]]
    data_rows.sort(key=lambda x: x[sort_col_index] if x[sort_col_index] else "")
    
    ws.delete_rows(2, ws.max_row)
    for sorted_row in data_rows:
        ws.append(sorted_row)
        if date_format_col and sorted_row[date_format_col - 1]:
            ws.cell(row=ws.max_row, column=date_format_col).number_format = 'yyyy/m/d'

def get_schedule_data(target_year):
    """race_dataのスケジュールを読み込み、日付・会場マップを一括生成する"""
    race_data_file = f"{target_year}_race_data.xlsx"
    available_dates, venue_data_map, date_map, venue_map = [], {}, {}, {}
    if not os.path.exists(race_data_file):
        return available_dates, venue_data_map, date_map, venue_map

    try:
        wb = openpyxl.load_workbook(race_data_file, data_only=True, read_only=True)
        if "schedule" in wb.sheetnames:
            ws_sched = wb["schedule"]
            for row in ws_sched.iter_rows(min_row=2, values_only=True):
                d_val = row[0]
                if d_val:
                    d_str = d_val.strftime('%Y-%m-%d') if isinstance(d_val, datetime) else str(d_val)[:10]
                    available_dates.append(d_str)
                    v_info = {}
                    for i in [1, 4, 7]:
                        if i + 2 < len(row) and row[i+1]:
                            v_id, v_name, v_day = row[i], row[i+1], row[i+2]
                            v_info[str(v_name)] = {"id": v_id, "day": v_day}
                            label = f"{v_id}回{v_name}{v_day}日"
                            date_map[label] = d_str
                            venue_map[label] = f"{v_id}_{v_name}"
                    venue_data_map[d_str] = v_info
    except Exception:
        pass
    return available_dates, venue_data_map, date_map, venue_map

def get_race_info_from_sheet(ws_race, search_text, target_r_num=None):
    """指定されたシートから対象レース群の詳細情報を抽出する"""
    races_info = {}
    target_col_idx = next((cell.column for cell in ws_race[1] if cell.value == search_text), None)
    if not target_col_idx: return races_info

    data_col_idx = target_col_idx 
    label_col_idx = target_col_idx - 1 

    last_row = 3
    for r in range(ws_race.max_row, 2, -1):
        if ws_race.cell(row=r, column=data_col_idx + 1).value is not None:
            last_row = r
            break
            
    all_rows = list(ws_race.iter_rows(min_row=3, max_row=last_row))
    race_start_indices = {}

    for i, row_cells in enumerate(all_rows):
        if label_col_idx < len(row_cells):
            val = str(row_cells[label_col_idx].value or "")
            match = re.search(r'(\d+)', val)
            if match:
                r_num = int(match.group(1))
                if 1 <= r_num <= 12 and r_num not in race_start_indices:
                    race_start_indices[r_num] = i

    sorted_races = sorted(race_start_indices.items())
    for idx, (r_num, start_idx) in enumerate(sorted_races):
        if target_r_num and r_num != target_r_num: continue
        
        end_idx = sorted_races[idx + 1][1] - 1 if idx + 1 < len(sorted_races) else len(all_rows) - 1
        
        race_col_data = [str(all_rows[i][data_col_idx].value or "") for i in range(start_idx, end_idx + 1)]

        races_info[r_num] = {
            'time': format_excel_time(all_rows[start_idx][data_col_idx + 1].value),
            'num': all_rows[start_idx][label_col_idx].value,
            'name': extract_race_name(race_col_data),
            'condition': extract_race_condition(race_col_data),
            'course': extract_race_course(race_col_data)
        }
    return races_info

# ==============================================================================
# ルーティング (Controllers)
# ==============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        with sqlite3.connect(DB_NAME) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT password FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()
            
            if user and check_password_hash(user[0], password):
                session.clear()
                session['login'] = True
                session['username'] = username
                return redirect(url_for('index'))
            else:
                flash('ユーザー名またはパスワードが違います。')
                
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('ログアウトしました。')
    return redirect(url_for('login'))

@app.route('/')
def index():
    stable_filter = request.args.get('stable', '')
    results = get_all_horses()
    if stable_filter:
        results = [h for h in results if h[6] == stable_filter]
    return render_template('index.html', results=results, stables=get_stables_list(), current_year=datetime.now().year)

@app.route('/add_horse_page')
@login_required
def add_horse_page():
    return render_template('add_horse.html', stables=get_stables_list(), default_year=datetime.now().year - 3)

@app.route('/add_horse', methods=['POST'])
@login_required
def add_horse():
    wb = openpyxl.load_workbook(horse_data)
    ws = wb["Horses"]
    name = request.form.get('name')
    birth_date = date(int(request.form.get('year')), int(request.form.get('month')), int(request.form.get('day')))

    if any(row[0] == name for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)):
        flash(f"エラー：『{name}』は既に登録されています。")
        return redirect('/add_horse_page')
        
    ws.append([name, request.form.get('gender'), birth_date, request.form.get('sire'), request.form.get('dam'), request.form.get('area'), request.form.get('stable_name')])
    sort_and_resize_table(ws, sort_col_index=0, date_format_col=3)
    wb.save(horse_data)
    return redirect(f"/horse/{name}")

@app.route('/add_stable', methods=['POST'])
@login_required
def add_stable():
    name, kana, area = request.form.get('stable_name'), request.form.get('kana'), request.form.get('area')
    if name and area and area in ["美浦", "栗東"]:
        wb = openpyxl.load_workbook(horse_data)
        sheet = wb[area]
        if name in [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]:
            flash(f"エラー: {name}厩舎は既に{area}に登録されています。")
            return redirect('/add_horse_page')

        sheet.append([name, kana_to_hira(re.sub(r'\s+', '', kana))])
        sort_and_resize_table(sheet, sort_col_index=1)
        wb.save(horse_data)   
    return redirect('/add_horse_page')

@app.route('/add_parent', methods=['GET', 'POST'])
@login_required
def add_parent():
    if request.method == 'POST':
        origin, p_type, p_name = request.form.get('origin'), request.form.get('p_type'), request.form.get('p_name')
        try:
            wb = openpyxl.load_workbook(horse_data)
            ws = wb[p_type] if p_type in wb.sheetnames else wb.create_sheet(title=p_type)
            if ws.max_row == 1 and not ws.cell(row=1, column=1).value: ws.append(["馬名", "生年月日", "父", "母"])

            y, m, d = request.form.get('year'), request.form.get('month'), request.form.get('day')
            birth_date = date(int(y), int(m), int(d)) if y and m and d else None

            found = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value == p_name:
                    row[1].value, row[2].value, row[3].value = birth_date, request.form.get('sire'), request.form.get('dam')
                    found = True
                    break
            
            if not found: ws.append([p_name, birth_date, request.form.get('sire'), request.form.get('dam')])
            sort_and_resize_table(ws, sort_col_index=0, date_format_col=2)
            wb.save(horse_data)
        except Exception as e:
            flash(f"エラーが発生しました: {e}")
        return redirect(f"/horse/{origin}") if origin else redirect('/')

    # GET
    p_type, p_name = request.args.get('type', 'Sire'), request.args.get('name', '')
    existing_data = {"year": "", "month": "", "day": "", "sire": "", "dam": ""}
    try:
        wb = openpyxl.load_workbook(horse_data, data_only=True)
        if p_type in wb.sheetnames:
            row = next((r for r in wb[p_type].iter_rows(values_only=True) if r[0] == p_name), None)
            if row:
                if row[1] and isinstance(row[1], (date, datetime)):
                    existing_data.update({"year": row[1].year, "month": row[1].month, "day": row[1].day})
                existing_data.update({"sire": row[2] or "", "dam": row[3] or ""})
    except: pass
    return render_template('add_parent.html', p_type=p_type, p_name=p_name, origin=request.args.get('origin', ''), data=existing_data)

@app.route('/update_horse', methods=['POST'])
@login_required
def update_horse():
    try:
        new_name = request.form.get('name')
        birth_date = date(int(request.form.get('year')), int(request.form.get('month')), int(request.form.get('day')))
        
        wb = openpyxl.load_workbook(horse_data)
        for row in wb["Horses"].iter_rows(min_row=2):
            if row[0].value == request.form.get('old_name'):
                row[0].value, row[1].value, row[2].value = new_name, request.form.get('gender'), birth_date
                row[3].value, row[4].value = request.form.get('sire'), request.form.get('dam')
                row[5].value, row[6].value = request.form.get('location'), request.form.get('stable_name')
                break
        wb.save(horse_data)
        return redirect(f"/horse/{new_name}")
    except Exception as e:
        return f"エラーが発生しました: {e}", 400

@app.route('/horse/<name>')
def horse_detail(name):
    all_horses = get_all_horses()
    horse = next((h for h in all_horses if h[0] == name), None)
    horse_races, schedule_cache = [], {}

    for f_path in glob.glob('*_entry_*.xlsx'):
        match = re.search(r'^(\d{4})_entry_(.+)\.xlsx$', os.path.basename(f_path))
        if not match: continue
        target_year, venue = match.group(1), match.group(2)

        if target_year not in schedule_cache:
            _, _, y_date_map, y_venue_map = get_schedule_data(target_year)
            schedule_cache[target_year] = (y_date_map, y_venue_map)
        
        date_map, venue_map = schedule_cache[target_year]
        wb_e = openpyxl.load_workbook(f_path, data_only=True)
        
        # race_dataを必要になった時だけ開く
        race_master_wb = None
        if os.path.exists(f"{target_year}_race_data.xlsx"):
            race_master_wb = openpyxl.load_workbook(f"{target_year}_race_data.xlsx", data_only=True, read_only=True)

        for sheet_name in wb_e.sheetnames:
            ws_e = wb_e[sheet_name]
            race_date = date_map.get(sheet_name, sheet_name)
            
            display_date = race_date
            try: display_date = f"{datetime.strptime(race_date, '%Y-%m-%d').strftime('%Y年%m月%d日')}({['月','火','水','木','金','土','日'][datetime.strptime(race_date, '%Y-%m-%d').weekday()]})"
            except: pass

            race_info_cache = {}
            if race_master_wb:
                m_sheet_name = venue_map.get(sheet_name)
                if m_sheet_name and m_sheet_name in race_master_wb.sheetnames:
                    race_info_cache = get_race_info_from_sheet(race_master_wb[m_sheet_name], sheet_name)

            for r_num in range(1, 13):
                rank_col, num_col, name_col, stat_col = (r_num - 1)*4 + 1, (r_num - 1)*4 + 2, (r_num - 1)*4 + 3, (r_num - 1)*4 + 4
                for row in range(3, ws_e.max_row + 1):
                    if ws_e.cell(row=row, column=name_col).value == name:
                        r_info = race_info_cache.get(r_num, {})
                        horse_races.append({
                            'sort_date': race_date, 'date_label': display_date, 'venue': venue, 'num': r_num,
                            'name': r_info.get('name', '-'), 'condition': r_info.get('condition', '-'), 'course': r_info.get('course', '-'),
                            'status': ws_e.cell(row=row, column=stat_col).value or "-", 'rank': ws_e.cell(row=row, column=rank_col).value or "-"
                        })

    horse_races.sort(key=lambda x: (x['sort_date'], x['num']), reverse=True)

    # 血統情報の取得
    sire_info, dam_info = {"line1": "不明", "line2": "不明"}, {"line3": "不明", "line4": "不明"}
    try:
        wb = openpyxl.load_workbook(horse_data, data_only=True)
        if horse and horse[3] and "Sire" in wb.sheetnames:
            s_row = next((r for r in wb["Sire"].iter_rows(values_only=True) if r[0] == horse[3]), None)
            if s_row: sire_info.update({"line1": s_row[2] or "不明", "line2": s_row[3] or "不明"})
        if horse and horse[4] and "Dam" in wb.sheetnames:
            d_row = next((r for r in wb["Dam"].iter_rows(values_only=True) if r[0] == horse[4]), None)
            if d_row: dam_info.update({"line3": d_row[2] or "不明", "line4": d_row[3] or "不明"})
    except: pass

    return render_template('horse_detail.html', horse=horse, horse_races=horse_races, sire_info=sire_info, dam_info=dam_info, current_year=datetime.now().year)

@app.route('/edit_horse/<name>')
@login_required
def edit_horse(name):
    horse = next((h for h in get_all_horses() if h[0] == name), None)
    return render_template('edit_horse.html', horse=horse, stables=get_stables_list(), current_year=datetime.now().year) if horse else ("Horse not found", 404)

@app.route('/races')
def race_list():
    req_date = request.args.get('date')
    target_year = req_date[:4] if req_date else str(datetime.now().year)
    
    available_dates, venue_data_map, _, _ = get_schedule_data(target_year)
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    future_dates = [d for d in available_dates if d >= today_str]
    date = req_date if req_date and req_date in available_dates else (future_dates[0] if future_dates else (available_dates[-1] if available_dates else today_str))

    current_day_venues = venue_data_map.get(date, {})
    default_venue = list(current_day_venues.keys())[0] if current_day_venues else ''
    venue = request.args.get('venue', default_venue)
    
    # 日付ブロック
    sorted_dates = sorted(available_dates)
    date_blocks, current_block = [], []
    for i, d in enumerate(sorted_dates):
        if i == 0: current_block.append(d)
        else:
            if (datetime.strptime(d, '%Y-%m-%d') - datetime.strptime(sorted_dates[i-1], '%Y-%m-%d')).days <= 2:
                current_block.append(d)
            else:
                date_blocks.append(current_block)
                current_block = [d]
    if current_block: date_blocks.append(current_block)

    current_date_block = next((b for b in date_blocks if date in b), [])
    display_dates = [{'value': d, 'label': f"{datetime.strptime(d, '%Y-%m-%d').month}/{datetime.strptime(d, '%Y-%m-%d').day}({['月','火','水','木','金','土','日'][datetime.strptime(d, '%Y-%m-%d').weekday()]})"} for d in current_date_block]

    day_races = {i: None for i in range(1, 13)}
    search_text = "開催情報が見つかりません"

    venue_info = current_day_venues.get(venue)
    # --- 修正箇所 ---
    if venue_info and os.path.exists(f"{target_year}_race_data.xlsx"):
        wb = openpyxl.load_workbook(f"{target_year}_race_data.xlsx", data_only=True, read_only=True)
        target_sheet_name = f"{venue_info['id']}_{venue}"
        if target_sheet_name in wb.sheetnames:
            search_text = f"{venue_info['id']}回{venue}{venue_info['day']}日"
            # 12レース分一括取得
            fetched_races = get_race_info_from_sheet(wb[target_sheet_name], search_text)
            
            # 重要：取得できたレースだけを上書きし、Noneにならないようにする
            for i in range(1, 13):
                if i in fetched_races:
                    day_races[i] = fetched_races[i]
                else:
                    # データが取れなかった場合も、HTMLで判定を通るように最低限の情報を入れる
                    day_races[i] = {
                        'num': f"{i}レース",
                        'name': '',
                        'condition': '情報なし',
                        'course': '',
                        'time': ''
                    }

    return render_template('race_list.html', date=date, venue=venue, day_races=day_races, available_venues=list(current_day_venues.keys()), available_dates=available_dates, display_dates=display_dates, search_text=search_text)

@app.route('/edit_race')
def race_detail():
    req_date, r_num_target = request.args.get('date'), request.args.get('num', default=1, type=int)
    target_year = req_date[:4] if req_date else str(datetime.now().year)
    
    available_dates, venue_data_map, _, _ = get_schedule_data(target_year)
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    future_dates = [d for d in available_dates if d >= today_str]
    date = req_date if req_date and req_date in available_dates else (future_dates[0] if future_dates else (available_dates[-1] if available_dates else today_str))

    current_day_venues = venue_data_map.get(date, {})
    default_venue = list(current_day_venues.keys())[0] if current_day_venues else ''
    venue = request.args.get('venue', default_venue)

    target_race_data, search_text = None, "開催情報が見つかりません"
    venue_info = current_day_venues.get(venue)
    
    if venue_info and os.path.exists(f"{target_year}_race_data.xlsx"):
        wb = openpyxl.load_workbook(f"{target_year}_race_data.xlsx", data_only=True, read_only=True)
        target_sheet_name = f"{venue_info['id']}_{venue}"
        if target_sheet_name in wb.sheetnames:
            search_text = f"{venue_info['id']}回{venue}{venue_info['day']}日"
            races_info = get_race_info_from_sheet(wb[target_sheet_name], search_text, target_r_num=r_num_target)
            target_race_data = races_info.get(r_num_target)

    available_horses_with_class = []
    if target_race_data:
        all_results_dict = load_all_horse_results()
        race_req_class = judge_required_class(target_race_data['condition'])
        
        for h in get_all_horses():
            horse_birthday = h[2]

            try:
                b_year = horse_birthday.year if isinstance(horse_birthday, datetime) else int(str(horse_birthday).split('/')[0])
                calculated_age = int(target_year) - b_year
            except:
                calculated_age = "不明"

            current_class = get_class_from_results(h[0], date, all_results_dict, horse_birthday)

            if race_req_class == "新馬":
                if current_class == "新馬":
                    available_horses_with_class.append({
                        'name': h[0],
                        'gender': h[1],
                        'age': calculated_age,
                        'class': current_class,
                        'stable': f"{h[5]}・{h[6]}"
                    })
            elif race_req_class == "未勝利":
                # 未勝利戦の場合は、新馬（初出走）も未勝利（既出走・勝星なし）も出られる
                if current_class in ["新馬", "未勝利"]:
                    available_horses_with_class.append({
                        'name': h[0],
                        'gender': h[1],
                        'age': calculated_age,
                        'class': current_class,
                        'stable': f"{h[5]}・{h[6]}"
                    })
            elif race_req_class == "オープン" or current_class == race_req_class:
                # その他のクラス
                available_horses_with_class.append({
                    'name': h[0],
                    'gender': h[1],
                    'age': calculated_age,
                    'class': current_class,
                    'stable': f"{h[5]}・{h[6]}"
                })
                
    entered_horses = []
    entry_file = f'{target_year}_entry_{venue}.xlsx'
    if os.path.exists(entry_file):
        wb_entry = openpyxl.load_workbook(entry_file, data_only=True)
        if search_text in wb_entry.sheetnames:
            ws_entry = wb_entry[search_text]
            rank_col, num_col, name_col, status_col = (r_num_target - 1)*4 + 1, (r_num_target - 1)*4 + 2, (r_num_target - 1)*4 + 3, (r_num_target - 1)*4 + 4
            for row in range(3, ws_entry.max_row + 1):
                name = ws_entry.cell(row=row, column=name_col).value
                if name:
                    entered_horses.append({
                        'rank': ws_entry.cell(row=row, column=rank_col).value or "",
                        'num': ws_entry.cell(row=row, column=num_col).value or "",
                        'name': name, 
                        'status': ws_entry.cell(row=row, column=status_col).value
                    })

    return render_template('race_detail.html',
                           date=date, venue=venue,
                           race=target_race_data,
                           available_dates=available_dates,
                           available_horses=available_horses_with_class,
                           all_horses=get_all_horses(),
                           entered_horses=entered_horses,
                           search_text=search_text)

@app.route('/save_entry', methods=['POST'])
@login_required
def save_entry():
    data = request.json
    target_year = data.get('date', datetime.now().strftime('%Y-%m-%d'))[:4]
    race_num, horse_name, entry_type, sheet_name = int(data.get('race_num')), data.get('horse_name'), data.get('entry_type'), data.get('sheet_name')
    horse_num, horse_rank = data.get('horse_num'), data.get('horse_rank')

    status_label = {"estimated": "想定", "special": "特別", "final": "確定"}.get(entry_type, "想定")
    file_name = f'{target_year}_entry_{data.get("venue")}.xlsx'
    
    wb = openpyxl.load_workbook(file_name) if os.path.exists(file_name) else openpyxl.Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames: ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        for r in range(1, 13):
            c = (r - 1) * 4 + 1
            ws.cell(row=1, column=c, value=f"{r}R")
            ws.cell(row=2, column=c, value="着順"); ws.cell(row=2, column=c+1, value="馬番")
            ws.cell(row=2, column=c+2, value="馬名"); ws.cell(row=2, column=c+3, value="出走")

    rank_col, num_col, name_col, status_col = (race_num - 1)*4 + 1, (race_num - 1)*4 + 2, (race_num - 1)*4 + 3, (race_num - 1)*4 + 4
    
    target_row = 3
    while ws.cell(row=target_row, column=name_col).value is not None and ws.cell(row=target_row, column=name_col).value != horse_name:
        target_row += 1
    
    ws.cell(row=target_row, column=name_col, value=horse_name)
    ws.cell(row=target_row, column=status_col, value=status_label)
    if horse_num: ws.cell(row=target_row, column=num_col, value=int(horse_num))
    if horse_rank: ws.cell(row=target_row, column=rank_col, value=int(horse_rank))

    wb.save(file_name)
    return {"status": "success"}, 200

if __name__ == '__main__':
    app.run(debug=True, port=5001 ,use_reloader=False)